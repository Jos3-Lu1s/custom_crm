# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError
from io import BytesIO
import base64
import xlsxwriter
from openpyxl import load_workbook

class CrmLeadWizard(models.TransientModel):
    # Wizard para exportar o importar oportunidades de CRM.
    _name = 'crm.lead.wizard'
    _description = 'Wizard para reintentar oportunidades de ventas perdidas'

    fecha_inicio = fields.Date(string='Fecha de inicio')
    fecha_fin = fields.Date(string='Fecha de fin')

    modo_operacion = fields.Selection([
        ('exportar', 'Exportar'),
        ('importar', 'Importar')
    ], string="Modo", default='exportar')

    archivo = fields.Binary(string="Subir archivo")

    def action_generar(self):
        # Ejecuta la acción del wizard según el modo (exportar o importar).
        if self.modo_operacion == 'exportar':
            stage_id = int(self.env['ir.config_parameter'].sudo().get_param('crm.lost_stage_id', default=0))
            stage = self.env['crm.stage'].browse(stage_id)

            domain = [
                ('type', '=', 'opportunity'),
            ]

            # Filtrar por etapa si existe

            if stage:
                domain.append(('stage_id', '=', stage.id))
            
            # Filtrar por fecha de inicio/fin si se especifica

            if self.fecha_inicio:
                domain.append(('date_last_stage_update', '>=', self.fecha_inicio))

            if self.fecha_fin:
                domain.append(('date_last_stage_update', '<=', self.fecha_fin))

            oportunidades = self.env['crm.lead'].search(domain)
            
            return self.crear_excel(oportunidades)
        
        elif self.modo_operacion == 'importar':
            catalogo = self.leer_excel()
            ids_a_perder = catalogo.get('no', [])
            ids_a_reintentar = catalogo.get('si', [])
            self.marcar_perdidas(ids_a_perder)
            self.duplicar_leads(ids_a_reintentar)
            return {
                'type': 'ir.actions.client',
                'tag': 'reload',
            }

    def crear_excel(self, oportunidades):
        """
        Genera un archivo Excel con las oportunidades dadas.
        :param oportunidades: recordset de crm.lead
        :return: acción Odoo para descargar el Excel
        """

        # Crear Excel en memoria
        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Oportunidades')

        # Formatos
        header_format = workbook.add_format({'bold': True})
        date_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})

        # Encabezados
        headers = [
            'Id',
            'Nombre',
            'Prima',
            'Cliente',
            'Vendedor',
            'Area',
            'Etapa',
            'Fecha vencimiento',
            'Intentar de nuevo',
            'Nuevo vencimiento',
        ]

        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # Datos
        row = 1
        for lead in oportunidades:
            worksheet.write(row, 0, lead.id or '')
            worksheet.write(row, 1, lead.name or '')
            worksheet.write(row, 2, lead.expected_revenue or 0)
            worksheet.write(row, 3, lead.partner_id.name or '')
            worksheet.write(row, 4, lead.user_id.name or '')
            worksheet.write(row, 5, lead.area_id.name or '')
            worksheet.write(row, 6, lead.stage_id.name or '')
            worksheet.write(row, 7, lead.date_deadline, date_format)
            worksheet.write(row, 8, 'No')
            worksheet.write(row, 9, '')
            row += 1

        workbook.close()
        output.seek(0)

        # Crear adjunto
        attachment = self.env['ir.attachment'].create({
            'name': 'oportunidades.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        # Descargar
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

        
    def leer_excel(self):
        """
        Lee el Excel subido y clasifica registros a reintentar ('si') o perder ('no').
        :return: dict {'si': {id: fecha}, 'no': [id]}
        :raises UserError: si faltan columnas o fechas
        """
        if not self.archivo:
            raise UserError("Debes subir un archivo Excel")

        data = base64.b64decode(self.archivo)
        try:
            wb = load_workbook(filename=BytesIO(data), read_only=True)
        except Exception:
            raise UserError(
                "El archivo subido no es compatible. \n"
                "Por favor, asegúrese de subir un archivo Excel (.xlsx) correcto."
            )
        ws = wb.active

        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]

        if 'Id' not in headers:
            raise UserError("El archivo no contiene la columna 'Id'")
        if 'Intentar de nuevo' not in headers:
            raise UserError("Falta la columna 'Intentar de nuevo'")
        if 'Nuevo vencimiento' not in headers:
            raise UserError("Falta la columna 'Nuevo vencimiento'")

        id_index = headers.index('Id')
        intentar_index = headers.index('Intentar de nuevo')
        n_vencimiento_index = headers.index('Nuevo vencimiento')

        catalogo = {
            'si': {},
            'no': []
        }

        for row in ws.iter_rows(min_row=2):
            record_id = row[id_index].value
            intentar = row[intentar_index].value
            nuevo_vencimiento = row[n_vencimiento_index].value

            if not record_id:
                continue

            valor = str(intentar).strip().lower() if intentar else 'no'

            if valor == 'si':
                if not nuevo_vencimiento:
                    raise UserError(
                        f"El registro con Id {record_id} tiene marcado 'Intentar de nuevo', pero no se ha indicado un 'Nuevo vencimiento'"
                    )
                catalogo['si'][record_id] = nuevo_vencimiento
            else:
                catalogo['no'].append(record_id)

        return catalogo

    def marcar_perdidas(self, ids):
        """
        Marca como perdidas las oportunidades activas con etapa asignada y probabilidad > 0.
        :param ids: lista de IDs de crm.lead
        """
        oportunidades = self.env['crm.lead'].browse(ids) if ids else self.env['crm.lead']
    
        for lead in oportunidades:
            if lead.stage_id and lead.probability != 0 and lead.active:
                lead.action_set_lost()

    def duplicar_leads(self, ids):
        """
        Duplica leads y actualiza stage y fecha de vencimiento.
        :param ids: dict {id: nueva_fecha}
        :return: recordset de leads duplicados
        :raises UserError: si la etapa de reintento no existe
        """
        if not ids:
            return self.env['crm.lead']

        stage_id = int(self.env['ir.config_parameter'].sudo().get_param('crm.retry_stage_id', default=0))
        stage = self.env['crm.stage'].browse(stage_id)

        oportunidades = self.env['crm.lead'].browse(ids.keys())

        new_leads = self.env['crm.lead']

        if not stage.exists():
            raise UserError(
                "La etapa de reintento configurada no existe.\n"
                "Por favor, configure una etapa válida en la configuración del módulo CRM antes de continuar."
            )

        for lead in oportunidades:
            nueva_fecha = ids.get(lead.id)
            new_lead = lead.copy(default={
                'stage_id': stage.id,
                'date_deadline': nueva_fecha,
            })
            new_leads += new_lead

        self.marcar_perdidas(oportunidades.ids)

        return new_leads